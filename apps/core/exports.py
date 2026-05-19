"""
Универсальный модуль экспорта данных в Excel (.xlsx).
Используется всеми разделами: лицензии, cloud, DNS, интернет.
"""
import io
from datetime import date
from decimal import Decimal

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.utils import timezone


# ── Стили ────────────────────────────────────────────────────────────────────

HEADER_FILL   = PatternFill("solid", fgColor="1B3A6B")
HEADER_FONT   = Font(color="FFFFFF", bold=True, size=10)
TITLE_FONT    = Font(bold=True, size=13, color="1B3A6B")
SUBHDR_FILL   = PatternFill("solid", fgColor="E8EDF5")
SUBHDR_FONT   = Font(bold=True, size=10, color="1B3A6B")
EXPIRED_FILL  = PatternFill("solid", fgColor="FDECEA")
WARNING_FILL  = PatternFill("solid", fgColor="FFF8E1")
CENTER        = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT          = Alignment(horizontal="left",   vertical="center", wrap_text=True)
RIGHT         = Alignment(horizontal="right",  vertical="center")

THIN  = Side(style="thin",   color="D0D7E3")
MED   = Side(style="medium", color="1B3A6B")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TOP_BORDER  = Border(top=MED)

NUM_FMT_INT   = '#,##0'
NUM_FMT_DEC   = '#,##0.00'
NUM_FMT_DATE  = 'DD.MM.YYYY'


def _cell(ws, row, col, value, font=None, fill=None, align=None,
          border=None, number_format=None):
    c = ws.cell(row=row, column=col, value=value)
    if font:           c.font   = font
    if fill:           c.fill   = fill
    if align:          c.alignment = align
    if border:         c.border = border
    if number_format:  c.number_format = number_format
    return c


def _header_row(ws, row, columns):
    for col_idx, (label, width) in enumerate(columns, start=1):
        c = ws.cell(row=row, column=col_idx, value=label)
        c.font      = HEADER_FONT
        c.fill      = HEADER_FILL
        c.alignment = CENTER
        c.border    = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[row].height = 28


def _data_cell(ws, row, col, value, expired=False, warning=False,
               number_format=None, align=None):
    fill = EXPIRED_FILL if expired else (WARNING_FILL if warning else None)
    c = _cell(ws, row, col, value,
              font=Font(size=9),
              fill=fill,
              align=align or LEFT,
              border=THIN_BORDER,
              number_format=number_format)
    return c


def _title_block(ws, title, subtitle, total_cols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    c = ws.cell(row=1, column=1, value=title)
    c.font      = TITLE_FONT
    c.alignment = LEFT

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
    c2 = ws.cell(row=2, column=1,
                 value=f"{subtitle}   |   Сформировано: {timezone.now().strftime('%d.%m.%Y %H:%M')}")
    c2.font      = Font(size=9, color="888888")
    c2.alignment = LEFT
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 16


# ── Лицензии ─────────────────────────────────────────────────────────────────

LICENSES_COLS = [
    ("№",                  4),
    ("Приложение",        28),
    ("Вендор",            20),
    ("Категория",         14),
    ("Объект",            14),
    ("Тип лицензии",      16),
    ("Куплено",            9),
    ("Исп-ся",             9),
    ("Свободно",           9),
    ("Цена за ед.",       14),
    ("Валюта",             8),
    ("Итого",             16),
    ("Итого (USD)",       14),
    ("№ договора",        16),
    ("Дата покупки",      13),
    ("Дата истечения",    13),
    ("Дней до истечения", 13),
    ("Статус",            12),
    ("Примечания",        30),
]

STATUS_LABELS = {
    'active':        'Активна',
    'expiring':      'Истекает (90 дн.)',
    'expiring_soon': 'Истекает (30 дн.)',
    'expired':       'Истекла',
}


def build_licenses_sheet(ws, licenses_qs, usd_rate, site_label="Все объекты"):
    _title_block(ws, "Реестр лицензий", site_label, len(LICENSES_COLS))
    ws.freeze_panes = "A5"
    _header_row(ws, 4, LICENSES_COLS)

    row = 5
    for idx, lic in enumerate(licenses_qs.select_related('app', 'app__vendor', 'site'), start=1):
        expired = lic.status == 'expired'
        warning = lic.status in ('expiring_soon', 'expiring')

        price_usd = lic.get_price_usd(usd_rate)
        total_usd = (price_usd * lic.quantity_total) if price_usd and lic.quantity_total else None

        vals = [
            idx,
            lic.app.name,
            lic.app.vendor.name if lic.app.vendor else '',
            lic.app.get_category_display(),
            lic.site.name,
            lic.get_license_type_display(),
            lic.quantity_total,
            lic.quantity_used,
            lic.quantity_free,
            float(lic.price_per_unit) if lic.price_per_unit else None,
            lic.currency,
            float(lic.total_cost)     if lic.total_cost     else None,
            float(total_usd)          if total_usd          else None,
            lic.contract_number or '',
            lic.purchase_date,
            lic.expiry_date,
            lic.days_until_expiry,
            STATUS_LABELS.get(lic.status, lic.status),
            lic.notes or '',
        ]

        num_fmts = [
            None, None, None, None, None, None,
            NUM_FMT_INT, NUM_FMT_INT, NUM_FMT_INT,
            NUM_FMT_DEC, None, NUM_FMT_DEC, NUM_FMT_DEC,
            None, NUM_FMT_DATE, NUM_FMT_DATE, NUM_FMT_INT,
            None, None,
        ]

        for col_idx, (val, fmt) in enumerate(zip(vals, num_fmts), start=1):
            _data_cell(ws, row, col_idx, val,
                       expired=expired, warning=warning,
                       number_format=fmt,
                       align=CENTER if col_idx == 1 else None)
        row += 1

    ws.row_dimensions[row].height = 20
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    c = ws.cell(row=row, column=1, value="ИТОГО")
    c.font      = Font(bold=True, size=10)
    c.alignment = RIGHT
    c.border    = Border(top=MED, bottom=MED)

    for col_idx in range(1, len(LICENSES_COLS) + 1):
        ws.cell(row=row, column=col_idx).border = Border(top=MED)

    return ws


# ── Cloud ────────────────────────────────────────────────────────────────────

CLOUD_COLS = [
    ("№",                  4),
    ("Имя сервера",       24),
    ("Провайдер",         18),
    ("Тип",               12),
    ("Объект",            14),
    ("ОС",                16),
    ("CPU",               10),
    ("RAM (GB)",           9),
    ("Диск (GB)",          9),
    ("IP адрес",          15),
    ("Назначение",        28),
    ("Стоимость / мес",   14),
    ("Валюта",             8),
    ("Период оплаты",     13),
    ("Следующий платёж",  14),
    ("Статус",            10),
    ("Примечания",        30),
]


def build_cloud_sheet(ws, servers_qs, site_label="Все объекты"):
    _title_block(ws, "Облачные серверы", site_label, len(CLOUD_COLS))
    ws.freeze_panes = "A5"
    _header_row(ws, 4, CLOUD_COLS)

    row = 5
    for idx, s in enumerate(servers_qs.select_related('site', 'provider'), start=1):
        stopped = s.status == 'stopped'
        vals = [
            idx,
            s.name,
            s.provider.name if s.provider else '',
            s.get_server_type_display(),
            s.site.name,
            s.os or '',
            s.cpu or '',
            s.ram_gb,
            s.disk_gb,
            str(s.ip_address) if s.ip_address else '',
            s.purpose or '',
            float(s.cost) if s.cost else None,
            s.currency,
            s.get_billing_period_display(),
            s.next_payment,
            s.get_status_display(),
            s.notes or '',
        ]
        num_fmts = [
            None, None, None, None, None, None, None,
            NUM_FMT_INT, NUM_FMT_INT,
            None, None, NUM_FMT_DEC, None, None, NUM_FMT_DATE, None, None,
        ]
        for col_idx, (val, fmt) in enumerate(zip(vals, num_fmts), start=1):
            fill = PatternFill("solid", fgColor="F5F5F5") if stopped else None
            c = _data_cell(ws, row, col_idx, val, number_format=fmt,
                           align=CENTER if col_idx == 1 else None)
            if stopped and fill:
                c.fill = fill
        row += 1
    return ws


# ── DNS / Домены ─────────────────────────────────────────────────────────────

DNS_COLS = [
    ("№",                 4),
    ("Домен",            28),
    ("Регистратор",      18),
    ("Объект",           14),
    ("Дата регистрации", 13),
    ("Дата истечения",   13),
    ("Дней до истечения",13),
    ("Стоимость / год",  14),
    ("Валюта",            8),
    ("Авто-продление",   12),
    ("Примечания",       30),
]


def build_dns_sheet(ws, domains_qs, site_label="Все объекты"):
    _title_block(ws, "Реестр доменов", site_label, len(DNS_COLS))
    ws.freeze_panes = "A5"
    _header_row(ws, 4, DNS_COLS)

    today = date.today()
    row = 5
    for idx, d in enumerate(domains_qs.select_related('site', 'registrar'), start=1):
        days    = (d.expiry_date - today).days if d.expiry_date else None
        expired = days is not None and days <= 0
        warning = days is not None and 0 < days <= 30

        vals = [
            idx,
            d.name,
            d.registrar.name if d.registrar else '',
            d.site.name,
            d.registration_date,
            d.expiry_date,
            days,
            float(d.cost) if d.cost else None,
            d.currency,
            'Да' if d.auto_renewal else 'Нет',
            d.notes or '',
        ]
        num_fmts = [
            None, None, None, None,
            NUM_FMT_DATE, NUM_FMT_DATE, NUM_FMT_INT,
            NUM_FMT_DEC, None, None, None,
        ]
        for col_idx, (val, fmt) in enumerate(zip(vals, num_fmts), start=1):
            _data_cell(ws, row, col_idx, val,
                       expired=expired, warning=warning,
                       number_format=fmt,
                       align=CENTER if col_idx == 1 else None)
        row += 1
    return ws


# ── Интернет / ISP ───────────────────────────────────────────────────────────

ISP_COLS = [
    ("№",                4),
    ("Услуга",          24),
    ("Оператор",        18),
    ("Тип услуги",      16),
    ("Объект",          14),
    ("Тариф",           18),
    ("Скорость",        12),
    ("IP адрес",        15),
    ("№ договора",      14),
    ("Стоимость / мес", 14),
    ("Валюта",           8),
    ("Дата истечения",  13),
    ("Дней до истечения",13),
    ("Статус",          10),
    ("Примечания",      30),
]

ISP_STATUS = {
    'active':   'Активен',
    'inactive': 'Неактивен',
    'pending':  'Ожидание',
}


def build_isp_sheet(ws, contracts_qs, site_label="Все объекты"):
    _title_block(ws, "Интернет / Связь", site_label, len(ISP_COLS))
    ws.freeze_panes = "A5"
    _header_row(ws, 4, ISP_COLS)

    today = date.today()
    row = 5
    for idx, c in enumerate(
            contracts_qs.select_related('site', 'operator', 'service_type'), start=1):
        days    = (c.expiry_date - today).days if hasattr(c, 'expiry_date') and c.expiry_date else None
        expired = days is not None and days <= 0
        warning = days is not None and 0 < days <= 30

        vals = [
            idx,
            c.service_name,
            c.operator.name if c.operator else '',
            c.service_type.name if c.service_type else '',
            c.site.name,
            c.tariff or '',
            c.speed or '',
            str(c.ip_address) if c.ip_address else '',
            c.contract_number or '',
            float(c.cost) if c.cost else None,
            c.currency,
            getattr(c, 'expiry_date', None),
            days,
            ISP_STATUS.get(getattr(c, 'status', ''), getattr(c, 'status', '')),
            c.notes or '',
        ]
        num_fmts = [
            None, None, None, None, None, None, None, None, None,
            NUM_FMT_DEC, None, NUM_FMT_DATE, NUM_FMT_INT, None, None,
        ]
        for col_idx, (val, fmt) in enumerate(zip(vals, num_fmts), start=1):
            _data_cell(ws, row, col_idx, val,
                       expired=expired, warning=warning,
                       number_format=fmt,
                       align=CENTER if col_idx == 1 else None)
        row += 1
    return ws


# ── HTTP Response helper ──────────────────────────────────────────────────────

def workbook_response(wb, filename_prefix="export"):
    today    = timezone.now().strftime("%Y%m%d")
    filename = f"{filename_prefix}_{today}.xlsx"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    response = HttpResponse(
        buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response
