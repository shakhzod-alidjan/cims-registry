"""
Management command to import data from the corporate Excel registry.

Usage:
    python manage.py import_excel path/to/Corporate_BP_Registry.xlsx
    python manage.py import_excel path/to/file.xlsx --site HQ --dry-run
"""
import sys
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction


class Command(BaseCommand):
    help = 'Import licenses, ISP contracts and domains from Excel file'

    def add_arguments(self, parser):
        parser.add_argument('file', type=str, help='Path to Excel file')
        parser.add_argument('--site', type=str, default=None,
                            help='Default site code (e.g. HQ). If omitted tries to detect from data.')
        parser.add_argument('--dry-run', action='store_true',
                            help='Validate only, do not save to database')
        parser.add_argument('--sheet', type=str, default=None,
                            help='Import only a specific sheet name')

    def handle(self, *args, **options):
        try:
            import openpyxl
        except ImportError:
            raise CommandError('openpyxl is required: pip install openpyxl')

        filepath = options['file']
        dry_run  = options['dry_run']
        default_site_code = options.get('site')

        self.stdout.write(f'📂 Opening: {filepath}')
        try:
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        except FileNotFoundError:
            raise CommandError(f'File not found: {filepath}')

        self.stdout.write(f'📋 Sheets found: {", ".join(wb.sheetnames)}')

        stats = {'licenses': 0, 'isp': 0, 'domains': 0, 'errors': 0}

        with transaction.atomic():
            for sheet_name in wb.sheetnames:
                if options['sheet'] and sheet_name != options['sheet']:
                    continue

                ws = wb[sheet_name]
                name_lower = sheet_name.lower()

                if any(k in name_lower for k in ('лицензи', 'license', 'реестр')):
                    self.stdout.write(f'\n  → Importing licenses from sheet: {sheet_name}')
                    stats['licenses'] += self._import_licenses(ws, default_site_code, dry_run)

                elif any(k in name_lower for k in ('интернет', 'isp', 'internet', 'связь')):
                    self.stdout.write(f'\n  → Importing ISP contracts from sheet: {sheet_name}')
                    stats['isp'] += self._import_isp(ws, default_site_code, dry_run)

                elif any(k in name_lower for k in ('dns', 'домен', 'domain')):
                    self.stdout.write(f'\n  → Importing domains from sheet: {sheet_name}')
                    stats['domains'] += self._import_domains(ws, default_site_code, dry_run)

            if dry_run:
                self.stdout.write(self.style.WARNING('\n⚠️  DRY RUN — no changes saved'))
                transaction.set_rollback(True)

        self.stdout.write(self.style.SUCCESS(
            f'\n✅ Done! Licenses: {stats["licenses"]} | '
            f'ISP: {stats["isp"]} | Domains: {stats["domains"]} | '
            f'Errors: {stats["errors"]}'
        ))

    # ── helpers ──────────────────────────────────────────────

    def _get_or_create_site(self, name_or_code: str):
        from apps.core.models import Site
        if not name_or_code:
            return None
        site, created = Site.objects.get_or_create(
            code=name_or_code.lower().replace(' ', '_'),
            defaults={'name': name_or_code, 'color': '#0052CC'},
        )
        if created:
            self.stdout.write(f'    + Created site: {site.name}')
        return site

    def _cell(self, row, idx):
        """Safe cell value getter."""
        try:
            val = row[idx].value if hasattr(row[idx], 'value') else row[idx]
            return str(val).strip() if val is not None else ''
        except (IndexError, AttributeError):
            return ''

    def _import_licenses(self, ws, default_site_code, dry_run) -> int:
        from apps.licenses.models import Vendor, BusinessApp, License
        from apps.core.models import Site
        import datetime

        rows = list(ws.iter_rows(min_row=2))
        count = 0

        for row in rows:
            app_name  = self._cell(row, 0)
            if not app_name or app_name.startswith('#'):
                continue

            try:
                lic_type   = self._cell(row, 1) or 'Подписка'
                vendor_name = self._cell(row, 2)
                qty_total  = self._cell(row, 3)
                qty_used   = self._cell(row, 4)
                price      = self._cell(row, 5)
                expiry_raw = self._cell(row, 6)
                notes      = self._cell(row, 7)
                site_name  = self._cell(row, 8) or default_site_code or 'HQ'

                site = self._get_or_create_site(site_name)

                vendor = None
                if vendor_name:
                    vendor, _ = Vendor.objects.get_or_create(name=vendor_name)

                app, _ = BusinessApp.objects.get_or_create(
                    name=app_name,
                    defaults={'vendor': vendor, 'category': 'other'},
                )
                if site:
                    app.sites.add(site)

                # Map license type string to model choice
                lt_map = {
                    'именная': 'named', 'named': 'named',
                    'конкурентная': 'concurrent', 'concurrent': 'concurrent',
                    'подписка': 'subscription', 'subscription': 'subscription',
                    'корпоративная': 'corporate',
                }
                lt = lt_map.get(lic_type.lower(), 'subscription')

                # Parse expiry date
                expiry = None
                if expiry_raw:
                    for fmt in ('%d.%m.%Y', '%Y-%m-%d', '%m/%d/%Y'):
                        try:
                            expiry = datetime.datetime.strptime(expiry_raw, fmt).date()
                            break
                        except ValueError:
                            continue

                lic, created = License.objects.update_or_create(
                    app=app, site=site, license_type=lt,
                    defaults={
                        'quantity_total': int(qty_total) if qty_total.isdigit() else None,
                        'quantity_used':  int(qty_used)  if qty_used.isdigit()  else None,
                        'price_per_unit': float(price)   if price and price.replace('.','').isdigit() else None,
                        'expiry_date':    expiry,
                        'notes':          notes,
                    }
                )
                count += 1
                mark = '+ Created' if created else '  Updated'
                self.stdout.write(f'    {mark}: {app_name} [{site}]')
            except Exception as e:
                self.stdout.write(self.style.ERROR(f'    ✗ Row error: {e}'))

        return count

    def _import_isp(self, ws, default_site_code, dry_run) -> int:
        from apps.internet.models import ISPOperator, ISPContract
        import datetime

        rows = list(ws.iter_rows(min_row=2))
        count = 0
        for row in rows:
            service_name = self._cell(row, 0)
            if not service_name or service_name.startswith('#'):
                continue
            try:
                op_name   = self._cell(row, 1)
                site_name = self._cell(row, 2) or default_site_code or 'HQ'
                tariff    = self._cell(row, 3)
                speed     = self._cell(row, 4)
                cost_uzs  = self._cell(row, 5)
                end_raw   = self._cell(row, 6)
                notes     = self._cell(row, 7)

                site = self._get_or_create_site(site_name)
                op   = None
                if op_name:
                    op, _ = ISPOperator.objects.get_or_create(name=op_name)

                end_date = None
                if end_raw:
                    for fmt in ('%d.%m.%Y', '%Y-%m-%d'):
                        try:
                            end_date = datetime.datetime.strptime(end_raw, fmt).date()
                            break
                        except ValueError:
                            continue

                ISPContract.objects.update_or_create(
                    service_name=service_name, site=site,
                    defaults={
                        'operator':   op,
                        'service_type': 'internet',
                        'tariff':      tariff,
                        'speed':       speed,
                        'cost_uzs':    float(cost_uzs.replace(' ','')) if cost_uzs else None,
                        'end_date':    end_date,
                        'notes':       notes,
                    }
                )
                count += 1
                self.stdout.write(f'    + ISP: {service_name} [{site}]')
            except Exception as e:
                self.stdout.write(self.style.ERROR(f'    ✗ {e}'))
        return count

    def _import_domains(self, ws, default_site_code, dry_run) -> int:
        from apps.dns.models import Registrar, Domain
        import datetime

        rows = list(ws.iter_rows(min_row=2))
        count = 0
        for row in rows:
            domain_name = self._cell(row, 0)
            if not domain_name or '.' not in domain_name:
                continue
            try:
                site_name  = self._cell(row, 1) or default_site_code or 'HQ'
                reg_name   = self._cell(row, 2)
                reg_raw    = self._cell(row, 3)
                exp_raw    = self._cell(row, 4)
                cost       = self._cell(row, 5)

                site = self._get_or_create_site(site_name)
                reg  = None
                if reg_name:
                    reg, _ = Registrar.objects.get_or_create(name=reg_name)

                def parse_date(s):
                    for fmt in ('%d.%m.%Y', '%Y-%m-%d'):
                        try:
                            return datetime.datetime.strptime(s, fmt).date()
                        except ValueError:
                            pass
                    return None

                Domain.objects.update_or_create(
                    name=domain_name,
                    defaults={
                        'site':              site,
                        'registrar':         reg,
                        'registration_date': parse_date(reg_raw),
                        'expiry_date':       parse_date(exp_raw),
                        'cost_usd':          float(cost) if cost and cost.replace('.','').isdigit() else None,
                    }
                )
                count += 1
                self.stdout.write(f'    + Domain: {domain_name} [{site}]')
            except Exception as e:
                self.stdout.write(self.style.ERROR(f'    ✗ {e}'))
        return count
